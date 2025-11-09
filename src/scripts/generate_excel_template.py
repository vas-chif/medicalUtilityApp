#!/usr/bin/env python3
"""
Generate Excel template for drug compatibility data entry

Usage:
    python generate_excel_template.py --output drug_template.xlsx
    
Requirements:
    pip install openpyxl pandas
"""

import argparse
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("⚠️  Excel libraries not installed. Install with:")
    print("   pip install openpyxl pandas")


def create_drug_template_excel(output_file: str):
    """Create Excel template for drug data entry"""
    
    if not HAS_EXCEL:
        print("❌ Excel libraries required. Install with: pip install openpyxl pandas")
        return
    
    # Create workbook
    wb = Workbook()
    
    # ============================================================
    # SHEET 1: DRUG LIST
    # ============================================================
    ws_drugs = wb.active
    ws_drugs.title = "Drug Database"
    
    # Headers
    headers = [
        'ID',
        'Drug Name',
        'Active Ingredient',
        'Category',
        'Concentration',
        'Route',
        'Requires CVC?',
        'Light Sensitive?',
        'Clinical Notes'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_drugs.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add example rows
    examples = [
        ['amiodarone', 'Amiodarone', 'Amiodarone Hydrochloride', 'antiarrhythmic', 
         '> 2mg/ml CVC+', 'intravenous', 'YES', 'YES', 
         'Class III antiarrhythmic - CVC required for concentrations > 2mg/ml. PROTECT FROM LIGHT'],
        ['dopamine', 'Dopamine', 'Dopamine Hydrochloride', 'vasopressor',
         '40mg/mL', 'intravenous', 'NO', 'YES',
         'Vasopressor - Cardiovascular support. Light sensitive.'],
        ['fentanyl', 'Fentanyl', 'Fentanyl Citrate', 'analgesic',
         '50mcg/mL', 'intravenous', 'NO', 'NO',
         'Opioid analgesic - Potent. Requires careful dosing.'],
    ]
    
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws_drugs.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Add 50 empty rows for data entry
    for row in range(5, 55):
        ws_drugs.cell(row=row, column=1).value = f"drug_{row-4:03d}"
    
    # Set column widths
    column_widths = [15, 20, 25, 18, 25, 15, 15, 15, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws_drugs.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze first row
    ws_drugs.freeze_panes = 'A2'
    
    # ============================================================
    # SHEET 2: COMPATIBILITY MATRIX
    # ============================================================
    ws_compat = wb.create_sheet("Compatibility Matrix")
    
    # Headers
    ws_compat['A1'] = 'Drug 1'
    ws_compat['B1'] = 'Drug 2'
    ws_compat['C1'] = 'Compatibility Code'
    ws_compat['D1'] = 'Description'
    
    # Style headers
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_compat[f'{col}1']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add legend
    legend_data = [
        ['Legend:', '', '', ''],
        ['C', 'Compatible', 'Green', 'Drugs can be mixed in the same solution'],
        ['Y', 'Y-site only', 'Yellow', 'Compatible only at Y-site, NOT mixed in bag'],
        ['I', 'Incompatible', 'Red', 'NEVER mix together (precipitation/degradation)'],
        ['!', 'Conflicting Data', 'Orange', 'Literature reports conflicting compatibility'],
        ['si', 'CVC Required', 'Light Green', 'Central Venous Catheter required'],
        ['', 'No Data', 'Grey', 'No compatibility data available'],
    ]
    
    for row_num, legend_row in enumerate(legend_data, 2):
        for col_num, value in enumerate(legend_row, 1):
            cell = ws_compat.cell(row=row_num, column=col_num)
            cell.value = value
            if row_num == 2:
                cell.font = Font(bold=True, size=11)
            if col_num == 1 and row_num > 2:
                cell.font = Font(bold=True)
    
    # Add color coding to legend
    color_fills = {
        3: PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),  # Green
        4: PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),  # Yellow
        5: PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),  # Red
        6: PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),  # Orange
        7: PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid"),  # Light Green
        8: PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid"),  # Grey
    }
    
    for row_num, fill in color_fills.items():
        ws_compat.cell(row=row_num, column=3).fill = fill
    
    # Add example compatibility entries
    example_compat = [
        ['Amiodarone', 'Dobutamine', 'I', 'Incompatible - Do not mix'],
        ['Amiodarone', 'Fentanyl', 'C', 'Compatible - Can be mixed'],
        ['Dopamine', 'Norepinephrine', '!', 'Conflicting data - Verify with pharmacist'],
    ]
    
    start_row = 11
    for row_num, compat_row in enumerate(example_compat, start_row):
        for col_num, value in enumerate(compat_row, 1):
            cell = ws_compat.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add 50 empty rows
    for row in range(start_row + 5, start_row + 55):
        pass  # Excel auto-creates rows
    
    # Set column widths
    ws_compat.column_dimensions['A'].width = 25
    ws_compat.column_dimensions['B'].width = 25
    ws_compat.column_dimensions['C'].width = 18
    ws_compat.column_dimensions['D'].width = 50
    
    # Freeze first row
    ws_compat.freeze_panes = 'A2'
    
    # ============================================================
    # SHEET 3: INSTRUCTIONS
    # ============================================================
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions_text = [
        ['🏥 DRUG COMPATIBILITY DATA ENTRY - INSTRUCTIONS', '', ''],
        ['', '', ''],
        ['Source:', 'Tabella Compatibilità Farmaci - Azienda Ospedaliera San Gerardo, Monza', ''],
        ['Authors:', 'Daniele Moretta, Alessia Gazzola, Stefania Poddaş, Alberto Lucchini', ''],
        ['Reference:', 'Prof. Antonio Pesenti (Università degli Studi di Milano-Bicocca)', ''],
        ['', '', ''],
        ['HOW TO USE THIS TEMPLATE:', '', ''],
        ['', '', ''],
        ['1. DRUG DATABASE SHEET:', '', ''],
        ['   - Fill in drug information from the San Gerardo Hospital table', '', ''],
        ['   - ID: Use lowercase with hyphens (e.g., "amiodarone")', '', ''],
        ['   - Drug Name: Full drug name (e.g., "Amiodarone")', '', ''],
        ['   - Active Ingredient: Chemical name if different', '', ''],
        ['   - Category: Choose from list (antibiotic, analgesic, cardiovascular, etc.)', '', ''],
        ['   - Concentration: From table (e.g., "> 2mg/ml CVC+")', '', ''],
        ['   - Requires CVC?: YES or NO', '', ''],
        ['   - Light Sensitive?: YES or NO (if "CONSERVARE AL RIPARO DALLA LUCE")', '', ''],
        ['   - Clinical Notes: Important clinical information', '', ''],
        ['', '', ''],
        ['2. COMPATIBILITY MATRIX SHEET:', '', ''],
        ['   - Enter drug pair compatibility codes', '', ''],
        ['   - Use codes from legend: C, Y, I, !, si, or leave empty', '', ''],
        ['   - Matrix is symmetric (A+B = B+A)', '', ''],
        ['', '', ''],
        ['3. EXPORT TO JSON:', '', ''],
        ['   - Save this file as CSV', '', ''],
        ['   - Use provided Python script to convert to JSON:', '', ''],
        ['     python convert_excel_to_json.py --input drug_data.xlsx --output drugs.json', '', ''],
        ['', '', ''],
        ['4. IMPORT TO PROJECT:', '', ''],
        ['   - Copy JSON data to src/data/drugs.ts', '', ''],
        ['   - Follow TypeScript format in existing file', '', ''],
        ['', '', ''],
        ['CATEGORIES:', '', ''],
        ['  - antibiotic', '', ''],
        ['  - analgesic', '', ''],
        ['  - cardiovascular', '', ''],
        ['  - anticoagulant', '', ''],
        ['  - sedative', '', ''],
        ['  - electrolyte', '', ''],
        ['  - vasopressor', '', ''],
        ['  - insulin', '', ''],
        ['  - diuretic', '', ''],
        ['  - antiarrhythmic', '', ''],
        ['  - other', '', ''],
    ]
    
    for row_num, text_row in enumerate(instructions_text, 1):
        for col_num, text in enumerate(text_row, 1):
            cell = ws_instructions.cell(row=row_num, column=col_num)
            cell.value = text
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="667EEA")
            elif text.startswith(('1.', '2.', '3.', '4.', 'HOW TO USE', 'CATEGORIES:')):
                cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws_instructions.column_dimensions['A'].width = 80
    ws_instructions.column_dimensions['B'].width = 60
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Excel template created: {output_file}")
    print(f"📊 Sheets: Drug Database, Compatibility Matrix, Instructions")
    print(f"🚀 Open in Excel/LibreOffice and start entering data!")


def main():
    parser = argparse.ArgumentParser(
        description='Generate Excel template for drug compatibility data'
    )
    parser.add_argument(
        '--output',
        '-o',
        type=str,
        default='drug_compatibility_template.xlsx',
        help='Output Excel file path (default: drug_compatibility_template.xlsx)'
    )
    
    args = parser.parse_args()
    create_drug_template_excel(args.output)


if __name__ == '__main__':
    main()
